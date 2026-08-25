import csv
import io

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.models.activity_event import ActivityEvent

COLUMNS = ["timestamp_utc", "event_type", "details", "received_at"]


def _format_details(details: dict | None) -> str:
    if not details:
        return ""
    return "; ".join(f"{k}={v}" for k, v in details.items())


def _row(e: ActivityEvent) -> list[str]:
    return [e.timestamp_utc.isoformat(), e.event_type, _format_details(e.details), e.received_at.isoformat()]


def export_events_csv(events: list[ActivityEvent]) -> io.BytesIO:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(COLUMNS)
    for e in events:
        writer.writerow(_row(e))
    return io.BytesIO(buffer.getvalue().encode("utf-8"))


def export_events_xlsx(events: list[ActivityEvent]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Activity Events"
    ws.append(COLUMNS)
    for e in events:
        ws.append(_row(e))

    for i, column in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(15, len(column) + 2)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
